import os
import re
import asyncio
import easyocr
import logging
import vt
import PyPDF2
import google.generativeai as genai
from docx import Document
from openpyxl import load_workbook
from datetime import datetime
from dotenv import load_dotenv
from pyrogram import Client, filters, enums
from pyrogram.types import (Message, InlineKeyboardMarkup, InlineKeyboardButton, 
                            ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove, CallbackQuery)
from pyrogram.errors import FloodWait, RPCError, PeerIdInvalid, ChatWriteForbidden

# Ma'lumotlar bazasi modulingiz
import database as db

# --- 1. SOZLAMALAR VA LOGGING ---
load_dotenv()
db.init_db()
logging.basicConfig(level=logging.INFO)

# OCR (MacBook CPU uchun)
try:
    reader = easyocr.Reader(['uz', 'en'], gpu=False)
except Exception as e:
    logging.error(f"OCR yuklanishda xato: {e}")

try:
    API_ID = int(os.getenv("API_ID"))
    API_HASH = os.getenv("API_HASH")
    BOT_TOKEN = os.getenv("BOT_TOKEN")
    SUPER_ADMIN = int(os.getenv("SUPER_ADMIN_ID", "0"))
    DATABASE_CHANNEL = int(os.getenv("DATABASE_CHANNEL", "0"))
    VT_API_KEY = os.getenv("VT_API_KEY")
    GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
except Exception as e:
    exit(f"❌ .env sozlamalarida xatolik: {e}")

genai.configure(api_key=GEMINI_API_KEY)
ai_model = genai.GenerativeModel('gemini-1.5-flash')

app = Client("ShieldPro_Official_AI", api_id=API_ID, api_hash=API_HASH, bot_token=BOT_TOKEN)

# --- 2. XAVFSIZLIK FILTRLARI ---
uzb_series = r"(?:AA|AB|AC|AD|AE|FA|KA|RR|UZ|TT|AF|BA)"
PATTERNS = {
    '💳 Bank Kartasi': r'(?:8600|9860|4444|5100|5300|6262|5445|5555)[\s-]?\d{4}[\s-]?\d{4}[\s-]?\d{4,8}',
    '🛂 Passport/ID': rf'{uzb_series}[\s-]?\d{{7}}', 
    '🆔 JSHSHIR': r'[3-6]\d{13}', 
    '🚗 Prava/Tex-pasport': r'\d{2}[\s-]?[A-Z]{2}[\s-]?\d{6}'
}

# --- 3. SCANNER FUNKSIYALARI ---

async def check_malicious_ai(text):
    if not GEMINI_API_KEY or len(text) < 10: return False
    try:
        prompt = f"Analyze for DLP. Reply 'DANGER' if contains Passport, JSHSHIR, Card or Phishing, else 'SAFE':\n\n{text[:1000]}"
        response = await asyncio.to_thread(ai_model.generate_content, prompt)
        return "DANGER" in response.text.upper()
    except: return False

async def vt_scan_file(file_path):
    if not VT_API_KEY: return False
    try:
        async with vt.Client(VT_API_KEY) as client:
            with open(file_path, "rb") as f:
                analysis = await client.scan_file_async(f)
                while True:
                    result = await client.get_object_async(f"/analyses/{analysis.id}")
                    if result.status == "completed": return result.stats['malicious'] > 0
                    await asyncio.sleep(5)
    except: return False

def extract_text_from_file(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    text = ""
    try:
        if ext == '.pdf':
            pdf = PyPDF2.PdfReader(file_path)
            for page in pdf.pages: text += page.extract_text() or ""
        elif ext in ['.docx', '.doc']:
            doc = Document(file_path)
            text = "\n".join([p.text for p in doc.paragraphs])
        elif ext in ['.xlsx', '.xls']:
            wb = load_workbook(file_path, data_only=True)
            for sheet in wb.sheetnames:
                for row in wb[sheet].iter_rows(values_only=True):
                    text += " ".join([str(cell) for cell in row if cell]) + " "
    except: pass
    return text

async def advanced_scan(message: Message):
    content = f"{message.text or ''} {message.caption or ''}".strip()
    clean_text = content.replace(" ", "").replace("-", "")
    
    for label, pattern in PATTERNS.items():
        if re.search(pattern, content, re.IGNORECASE) or re.search(pattern, clean_text): return label
    
    if await check_malicious_ai(content): return "⚠️ Shubhali mazmun (AI)"

    if message.document:
        if message.document.file_size > 20 * 1024 * 1024: return None
        path = await message.download()
        if await vt_scan_file(path):
            os.remove(path); return "🦠 Virus (Malware)"
        file_text = extract_text_from_file(path)
        os.remove(path)
        if file_text:
            for label, pattern in PATTERNS.items():
                if re.search(pattern, file_text.replace(" ",""), re.IGNORECASE): return f"{label} (Faylda)"
            if await check_malicious_ai(file_text): return "⚠️ Maxfiy ma'lumot (Faylda)"

    if message.photo:
        img_path = await message.download()
        loop = asyncio.get_event_loop()
        results = await loop.run_in_executor(None, reader.readtext, img_path)
        os.remove(img_path)
        detected_text = " ".join([res[1] for res in results]).replace(" ","")
        for label, pattern in PATTERNS.items():
            if re.search(pattern, detected_text, re.IGNORECASE): return f"{label} (Rasmda)"
    return None

# --- 4. INTERFEYS VA TUGMALAR ---

active_chats = {} 
user_states = {}

def get_main_menu(user_id):
    btns = [
        [KeyboardButton("🛡 Bot Imkoniyatlari")],
        [KeyboardButton("📊 Statistika"), KeyboardButton("👨‍💻 Admin bilan bog'lanish")]
    ]
    if user_id == SUPER_ADMIN:
        btns.append([KeyboardButton("⚙️ Admin Paneli")])
    return ReplyKeyboardMarkup(btns, resize_keyboard=True)

def get_admin_panel():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📊 Umumiy Stats", callback_data="admin_stats"), InlineKeyboardButton("🏢 Guruhlar", callback_data="admin_groups")],
        [InlineKeyboardButton("🚫 Oxirgi Xavflar", callback_data="admin_threats"), InlineKeyboardButton("📢 Global Xabar", callback_data="admin_broadcast")],
        [InlineKeyboardButton("❌ Paneldan chiqish", callback_data="admin_close")]
    ])

# --- 5. PRIVATE HANDLER (USER & ADMIN INTERFACE) ---

@app.on_message(filters.private, group=-1)
async def private_manager(client, message: Message):
    user_id = message.from_user.id
    text = message.text

    # Admin bilan jonli suhbat
    if user_id in active_chats:
        if text == "❌ Suhbatni yakunlash":
            partner = active_chats.pop(user_id)
            active_chats.pop(partner, None)
            await message.reply("🔚 Suhbat yakunlandi.", reply_markup=get_main_menu(user_id))
            await client.send_message(partner, "🔚 Suhbat yakunlandi.", reply_markup=get_main_menu(partner))
            return
        await message.copy(active_chats[user_id])
        return

    if text == "/start":
        user = db.get_user(user_id)
        if not user:
            user_states[user_id] = {"step": "wait_name"}
            await message.reply("👋 Xush kelibsiz! ShieldPro DLP tizimidan foydalanish uchun ismingizni kiriting:", reply_markup=ReplyKeyboardRemove())
        else:
            await message.reply("🛡 ShieldPro AI tizimi himoyaga tayyor!", reply_markup=get_main_menu(user_id))
        return

    # Ro'yxatdan o'tish
    state = user_states.get(user_id, {}).get("step")
    if state == "wait_name":
        user_states[user_id] = {"step": "wait_phone", "name": text}
        await message.reply("📞 Telefon raqamingizni yuboring:", reply_markup=ReplyKeyboardMarkup([[KeyboardButton("📞 Kontaktni ulashish", request_contact=True)]], resize_keyboard=True))
        return
    elif message.contact and state == "wait_phone":
        db.register_user(user_id, user_states[user_id]["name"], message.contact.phone_number)
        user_states.pop(user_id)
        await message.reply("✅ Ro'yxatdan o'tdingiz!", reply_markup=get_main_menu(user_id))
        return

    # Menu funksiyalari
    if text == "🛡 Bot Imkoniyatlari":
        info = ("🚀 **ShieldPro Ultra AI nimalarga qodir?**\n\n"
                "1️⃣ **DLP Nazorati:** Guruhlarda Passport, Karta va JSHSHIR ma'lumotlarini aniqlaydi.\n"
                "2️⃣ **OCR Scan:** Rasmlar ichidagi yozuvlarni o'qiydi.\n"
                "3️⃣ **Fayl Analiz:** PDF, Word, Excel ichidagi maxfiy ma'lumotlarni topadi.\n"
                "4️⃣ **Antivirus:** Zararli fayl va havolalarni bloklaydi.\n"
                "5️⃣ **AI Himoya:** Gemini AI orqali shubhali xabarlarni tahlil qiladi.")
        await message.reply(info, reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("➕ Guruhga qo'shish", url=f"https://t.me/{(await client.get_me()).username}?startgroup=new")]]))

    elif text == "📊 Statistika":
        u, g, t = db.get_stats()
        await message.reply(f"📊 **Hozirgi holat:**\n👤 Userlar: `{u}`\n🏢 Guruhlar: `{g}`\n🚫 Tahdidlar: `{t}`")

    elif text == "👨‍💻 Admin bilan bog'lanish":
        active_chats[user_id] = SUPER_ADMIN
        active_chats[SUPER_ADMIN] = user_id
        await message.reply("✍️ Xabaringizni yozing, admin tez orada javob beradi.", reply_markup=ReplyKeyboardMarkup([[KeyboardButton("❌ Suhbatni yakunlash")]], resize_keyboard=True))

    elif text == "⚙️ Admin Paneli" and user_id == SUPER_ADMIN:
        await message.reply("🛠 **Boshqaruv Paneli:**", reply_markup=get_admin_panel())

# --- 6. MONITORING HANDLER ---

@app.on_message((filters.group | filters.channel) & ~filters.service, group=1)
async def monitor_handler(client, message: Message):
    db.add_group(message.chat.id, message.chat.title)
    if message.from_user and message.from_user.id == SUPER_ADMIN: return

    threat = await advanced_scan(message)
    if threat:
        db.log_threat(message.chat.id, message.from_user.id or 0, threat)
        
        if DATABASE_CHANNEL != 0:
            try:
                log = (f"🛡 **DLP ALERT**\n🏢 Guruh: `{message.chat.title}`\n"
                       f"👤 User: {message.from_user.mention if message.from_user else 'N/A'}\n"
                       f"⚠️ Tahdid: **{threat}**")
                await client.send_message(DATABASE_CHANNEL, log)
            except: pass

        try:
            await message.delete()
            warn = await message.reply(f"❌ {message.from_user.mention}, xabaringizda **{threat}** aniqlandi!")
            await asyncio.sleep(5); await warn.delete()
        except: pass

# --- 7. CALLBACK HANDLER ---

@app.on_callback_query()
async def cb_handler(client, cb: CallbackQuery):
    if cb.from_user.id != SUPER_ADMIN: return
    
    u, g, t = db.get_stats()
    try:
        if cb.data == "admin_stats":
            await cb.message.edit_text(f"📊 **Statistika:**\n\nFoydalanuvchilar: {u}\nGuruhlar: {g}\nXavflar: {t}", reply_markup=get_admin_panel())
        elif cb.data == "admin_groups":
            await cb.message.edit_text(f"🏢 **Guruhlar soni:** {g}\nBot barcha guruhlarda faol nazorat olib bormoqda.", reply_markup=get_admin_panel())
        elif cb.data == "admin_threats":
            await cb.message.edit_text(f"🚫 **Jami qaytarilgan xavflar:** {t}\n\nTizim xavfsiz holatda. ✅", reply_markup=get_admin_panel())
        elif cb.data == "admin_close":
            await cb.message.delete()
    except RPCError: pass # "Message is not modified" xatosini chetlab o'tish

if __name__ == "__main__":
    print("🚀 ShieldPro Ultra AI ishga tushdi...")
    app.run()